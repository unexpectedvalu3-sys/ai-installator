# -*- coding: utf-8 -*-
"""Comparateur de garanties — MUTUELLE SANTÉ.

Plusieurs devis/offres assureurs (PDF) -> tableau comparatif aligné sur une grille
standard + synthèse client. Tourne sur **GLM 5.2 via OpenRouter** (bien moins cher
que Claude sur ce gros volume de texte) ; l'extraction de dossiers reste sur Sonnet.

Le texte des PDF est extrait LOCALEMENT (pdfplumber) puis envoyé comme texte au
LLM. On ne s'appuie plus sur le parseur pdf-text d'OpenRouter : il échoue sur les
PDF générés par Chrome/Skia (sans table ToUnicode) et renvoyait un document vide,
ce qui faisait disparaître le devis de la comparaison. Le parseur reste actif en
secours pour les PDF scannés (sans texte extractible) et les images.
"""
import os, json, base64, io
import httpx

import llm  # helpers LLM partagés (_parse_json, _call_llm) — voir llm.py

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

# Grille réelle du cabinet (comparatif complémentaire santé) — 5 modules A→E.
MODULES = [
    ("A", "Hospitalisation", [
        "Honoraires chirurgien/anesthésiste/réanimateur (OPTAM)",
        "Honoraires (Hors OPTAM)",
        "Frais de séjour (secteur conventionné)",
        "Frais de séjour (non conventionné)",
        "Chambre particulière — chirurgie (€/jour)",
        "Chambre particulière — rééducation/convalescence (€/jour)",
        "Chambre particulière — psychiatrie (€/jour)",
        "Lit accompagnant (€/jour)",
        "Frais annexes (TV, Wifi)",
    ]),
    ("B", "Soins courants", [
        "Consultations/visites généralistes & spécialistes (OPTAM)",
        "Consultations/visites (Non OPTAM)",
        "Actes techniques (OPTAM)",
        "Actes techniques (Non OPTAM)",
        "Radiologie (OPTAM)",
        "Radiologie (Non OPTAM)",
        "Auxiliaires médicaux, analyses & laboratoire",
        "Pharmacie prise en charge par l'AM",
        "Pharmacie non remboursée (avec ordonnance)",
    ]),
    ("C", "Dentaire", [
        "Soins dentaires",
        "Prothèses dentaires remboursées par l'AM",
        "Prothèses dentaires (Hors Réforme)",
        "Orthodontie prise en charge par l'AM",
        "Orthodontie non remboursée par l'AM",
        "Implant (racine + pilier, par an/bénéficiaire)",
        "Parodontologie (forfait par an/bénéficiaire)",
    ]),
    ("D", "Optique", [
        "Monture + verres simples",
        "Monture + verres complexes",
        "Lentilles (remboursées ou non par l'AM)",
        "Chirurgie réfractive de l'œil (par œil)",
    ]),
    ("E", "Prévention & autres", [
        "Prothèses auditives",
        "Bonus",
        "Consommables, piles & accessoires",
        "Autre appareillage",
        "Cures thermales adaptées par l'AM (par cure)",
        "Forfait annuel cure thermale",
        "Prime de naissance",
        "Médecine douce (ostéo, chiro, acupuncture, podologie…)",
    ]),
]


def _grid_text():
    lines = []
    for code, name, postes in MODULES:
        lines.append(f"{code} — {name} :")
        lines += [f"  - {p}" for p in postes]
    return "\n".join(lines)

PROMPT_TPL = """Tu es l'assistant d'un courtier en assurance. On te fournit PLUSIEURS devis
de MUTUELLE SANTÉ (un document PDF par assureur). Ta mission :

1) Identifie CHAQUE formule/offre présentée dans chaque document. Un seul PDF peut
   contenir PLUSIEURS formules (ex. un tableau comparatif de niveaux 1/2/3, ou des
   options). Tu DOIS toutes les extraire : une entrée dans "offres" par formule.
   Ne fusionne JAMAIS plusieurs formules en une seule. Ne saute JAMAIS une formule.
2) Pour CHAQUE offre, relève : assureur, nom de la formule/niveau, cotisation mensuelle TTC,
   et les FRAIS D'ADHÉSION s'ils figurent (ex. "30 €", "2 x 30 €", ou "—" si absent).
3) Pour CHAQUE offre, relève le niveau de remboursement de CHAQUE poste de la grille ci-dessous,
   EXACTEMENT comme écrit sur le devis (ex. "200% BR", "300 €", "100% FR", "forfait 150 €/an").
   Si un poste n'est pas couvert ou pas mentionné : "—".
4) Identifie le SOUSCRIPTEUR (civilité + nom) s'il figure sur un devis.
5) Détermine la RECOMMANDATION (meilleur rapport garanties/prix au vu des offres).

RÈGLE CRITIQUE — EXHAUSTIVITÉ :
- Relis CHAQUE document en ENTIER et liste TOUTES les formules qui y sont présentées.
- Si un document contient 3 formules (ex. Niveau 1, 2, 3), tu DOIS créer 3 offres.
- Compte bien : le nombre d'offres dans ta réponse doit correspondre au total de
  toutes les formules trouvées dans tous les documents. Si tu en as oublié, recompte.

Grille des postes, groupée par MODULE — utilise EXACTEMENT ces libellés, dans cet ordre :
{grid}

Réponds UNIQUEMENT par un JSON valide, sans texte autour :
{{
  "produit": "Mutuelle santé",
  "client": {{"civilite": "Monsieur|Madame|Mademoiselle (ou vide)", "nom": "nom du souscripteur (ou vide)"}},
  "offres": [ {{"assureur":"", "formule":"", "cotisation":"XX,XX €/mois", "frais_adhesion":"XX € ou —"}} ],
  "recommandation": {{"offre":"<assureur — formule>", "pourquoi":"1-2 phrases"}},
  "garanties": [ {{"module":"<code> — <nom du module>", "poste":"<libellé EXACT>", "valeurs":["offre 1","offre 2","..."]}} ],
  "warnings": ["info manquante / incertaine"]
}}

Règles STRICTES :
- "valeurs" est aligné sur l'ORDRE de "offres" (même nombre d'éléments).
- "garanties" couvre TOUS les postes de la grille, dans l'ordre, avec le bon "module".
- N'invente AUCUN chiffre. Poste absent = "—". Montants avec virgule décimale.
- **assureur** = l'organisme qui porte le risque / émet le contrat. Sur un devis,
  c'est l'entité citée dans les MENTIONS LÉGALES en bas de page (ex. « X SAS
  d'assurance / de courtage en assurances »), PAS le cabinet/conseiller en haut
  (« votre conseil », « réunion X assurances » = le courtier émetteur). N'invente
  JAMAIS un nom d'assureur qui n'apparaît pas dans le document : si le nom n'est
  pas identifiable, mets "" et ajoute un warning « Assureur à confirmer ».
"""


def _content_block(name, content):
    b64 = base64.standard_b64encode(content).decode()
    ext = name.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        return {"type": "file", "file": {"filename": name,
                                         "file_data": f"data:application/pdf;base64,{b64}"}}
    media = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp", "gif": "gif"}.get(ext, "jpeg")
    return {"type": "image_url", "image_url": {"url": f"data:image/{media};base64,{b64}"}}


def _extract_pdf_text(content):
    """Extrait le texte d'un PDF LOCALEMENT via pdfplumber. Renvoie le texte
    (str) ou '' si rien d'extractible (PDF scanné / image).

    Indispensable : le parseur pdf-text d'OpenRouter renvoie un document VIDE
    sur les PDF générés par Chrome/Skia (polices sans table ToUnicode), donc le
    devis n'était pas détecté par la comparaison. pdfplumber (pdfminer.six)
    exploite l'encoding des polices et récupère le texte, accents compris.
    """
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            parts = [p.extract_text() or "" for p in pdf.pages]
        return "\n".join(parts).strip()
    except Exception:
        return ""


def _build_blocks(files):
    """Construit les blocs de contenu envoyés au LLM.

    - PDF avec texte extractible : texte extrait localement, encapsulé dans un
      marqueur `===== DOCUMENT i : <nom> =====` pour que le LLM distingue chaque
      devis (un assureur/offre par document).
    - PDF sans texte (scanné) / images : envoyé brut (file/image_url) -> le
      parseur pdf-text d'OpenRouter ou la vision du modèle prennent le relais.
    """
    blocks, fallback = [], []
    for i, (name, content) in enumerate(files, start=1):
        if name.lower().rsplit(".", 1)[-1] == "pdf":
            txt = _extract_pdf_text(content)
            if txt:
                blocks.append({"type": "text",
                               "text": f"===== DOCUMENT {i} : {name} =====\n{txt}"})
            else:
                fallback.append((name, content))
        else:
            fallback.append((name, content))
    for name, content in fallback:
        blocks.append(_content_block(name, content))
    return blocks


def _call_openrouter(files, prompt):
    key = os.environ.get("OPENROUTER_API_KEY", "")
    model = os.environ.get("COMPARE_MODEL", "z-ai/glm-5.2")
    if not key:
        raise RuntimeError("OPENROUTER_API_KEY non configurée")
    content = _build_blocks(files) + [{"type": "text", "text": prompt}]
    body = {
        "model": model,
        "messages": [{"role": "user", "content": content}],
        "max_tokens": 16000,               # large : tableau 22 postes + synthèse, sans troncature
        "reasoning": {"enabled": False},   # GLM 5.2 raisonne sinon : on coupe -> JSON direct, moins cher
        # Parseur PDF texte gratuit d'OpenRouter — n'est plus utilisé pour les PDF
        # textuels (on envoie le texte extrait localement, plus fiable). Conservé en
        # secours pour les PDF scannés (sans texte extractible) envoyés comme fichiers.
        "plugins": [{"id": "file-parser", "pdf": {"engine": "pdf-text"}}],
    }
    r = httpx.post(OPENROUTER_URL, json=body, timeout=180, headers={
        "Authorization": f"Bearer {key}",
        "HTTP-Referer": "https://modulr-app",
        "X-Title": "Modulr Comparateur",
        "Content-Type": "application/json",
    })
    r.raise_for_status()
    j = r.json()
    choices = j.get("choices")
    if not choices:
        raise RuntimeError(f"OpenRouter: réponse sans choix ({str(j)[:200]})")
    out = choices[0]["message"].get("content")
    if not out:
        raise RuntimeError("OpenRouter: contenu vide")
    return out


SYNTHESE_PROMPT = """Tu rédiges, pour un courtier en assurance, un COURRIER de synthèse à envoyer
au client après comparaison de mutuelles santé. Respecte EXACTEMENT le style demandé.

Salutation : commence par « {bonjour} »
Offres comparées :
{offres}
Recommandation à défendre : {reco}
Repères de garanties (contexte — ne pas tout recopier) :
{garanties}

STYLE OBLIGATOIRE :
- CONCIS : environ 180-220 mots AU TOTAL. Quelques paragraphes courts. Ne dépasse pas.
- Phrase d'intro courte (ex. « Comme convenu, j'ai comparé pour vous trois mutuelles santé. Voici ma recommandation. »).
- LE CONSEIL D'ABORD : recommande clairement l'offre retenue (nom + cotisation mis en avant) et explique pourquoi en 2-3 arguments clés reliés aux besoins. NE liste PAS tous les postes de garanties. Surtout PAS un catalogue qui décrit chaque offre l'une après l'autre.
- Ensuite une courte section « Pour comparaison : » listant les AUTRES offres, UNE ligne chacune (un atout + une limite).
- ZÉRO jargon : ne JAMAIS écrire « BR », « base de remboursement », « % BR ». Traduis en clair (« bons remboursements », « couverture renforcée », « prise en charge élevée »).
- AUCUN markdown : pas de **gras**, pas de titres, pas de « --- ». Texte simple, en français courant.
- Ton professionnel, humain, chaleureux mais sobre. Vouvoiement.
- Une phrase rappelant que le client reste libre de son choix et ta disponibilité.
- Termine par une ligne « Cordialement, » (n'invente AUCUN nom de conseiller).
Réponds UNIQUEMENT par le texte du courrier, sans titre ni guillemets.
"""


def _synthese_client(data):
    """Rédige le courrier de synthèse sur Sonnet (prose FR plus fine que GLM)."""
    cl = data.get("client") or {}
    greet = (str(cl.get("civilite") or "").strip() + " " + str(cl.get("nom") or "").strip()).strip()
    bonjour = f"Bonjour {greet}," if greet else "Bonjour,"
    offres = "\n".join(f"- {o.get('assureur','')} — {o.get('formule','')} — {o.get('cotisation','')}"
                       for o in data.get("offres", [])) or "—"
    reco = data.get("recommandation") or {}
    gar = "\n".join(f"- {g.get('poste','')}: {' | '.join(g.get('valeurs', []))}"
                    for g in data.get("garanties", [])) or "—"
    prompt = SYNTHESE_PROMPT.format(bonjour=bonjour, offres=offres,
                                    reco=f"{reco.get('offre','')} — {reco.get('pourquoi','')}", garanties=gar)
    try:
        return llm._call_llm([], prompt, max_tokens=1200).strip()
    except Exception:
        return ""


def compare(files):
    """files = liste de (filename, bytes). Renvoie le dict de comparaison."""
    prompt = PROMPT_TPL.format(grid=_grid_text())
    data = llm._parse_json(_call_openrouter(files, prompt))
    for k, v in (("client", {}), ("offres", []), ("garanties", []), ("recommandation", {}), ("warnings", [])):
        data.setdefault(k, v)
    data["synthese_client"] = _synthese_client(data)   # toujours Sonnet (qualité de la prose)
    return data


def _norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", s or "")
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()


def _reco_index(data):
    """Même logique que recoIndex() côté JS (comparateur.html) : retrouve l'offre
    recommandée par correspondance floue sur assureur/formule, -1 si aucune."""
    offres = data.get("offres") or []
    cible = _norm((data.get("recommandation") or {}).get("offre"))
    if not cible:
        return -1
    for i, o in enumerate(offres):
        k = _norm(f"{o.get('assureur','')} {o.get('formule','')}")
        if cible in k or k in cible or _norm(o.get("assureur")) in cible:
            return i
    return -1


def build_xlsx(data):
    """Construit le classeur Excel du tableau comparatif (mêmes modules/lignes
    que l'affichage web, colonne recommandée surlignée). Renvoie des bytes .xlsx."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    offres = data.get("offres") or []
    garanties = data.get("garanties") or []
    ri = _reco_index(data)

    header_fill = PatternFill("solid", fgColor="FBEADF")
    reco_fill = PatternFill("solid", fgColor="C7E2DB")
    bold = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparatif"

    c = ws.cell(row=1, column=1, value="Garantie")
    c.font = bold
    c.fill = header_fill
    for i, o in enumerate(offres):
        c = ws.cell(row=1, column=2 + i, value=f"{o.get('assureur', '')} — {o.get('formule', '')}")
        c.font = bold
        c.fill = reco_fill if i == ri else header_fill
        c.alignment = wrap_top

    r = 2
    last_mod = None
    for g in garanties:
        mod = g.get("module") or "Garanties"
        if mod != last_mod:
            last_mod = mod
            c = ws.cell(row=r, column=1, value=mod)
            c.font = bold
            c.fill = header_fill
            for i in range(len(offres)):
                ws.cell(row=r, column=2 + i).fill = header_fill
            r += 1
        ws.cell(row=r, column=1, value=g.get("poste", ""))
        for i, v in enumerate(g.get("valeurs") or []):
            c = ws.cell(row=r, column=2 + i, value=v)
            if i == ri:
                c.fill = reco_fill
        r += 1

    c = ws.cell(row=r, column=1, value="Tarif")
    c.font = bold
    c.fill = header_fill
    for i in range(len(offres)):
        ws.cell(row=r, column=2 + i).fill = header_fill
    r += 1
    ws.cell(row=r, column=1, value="Frais d'adhésion")
    for i, o in enumerate(offres):
        c = ws.cell(row=r, column=2 + i, value=o.get("frais_adhesion", "—"))
        if i == ri:
            c.fill = reco_fill
    r += 1
    c = ws.cell(row=r, column=1, value="Cotisation mensuelle")
    c.font = bold
    for i, o in enumerate(offres):
        c = ws.cell(row=r, column=2 + i, value=o.get("cotisation", ""))
        c.font = bold
        if i == ri:
            c.fill = reco_fill

    ws.column_dimensions["A"].width = 44
    for i in range(len(offres)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 22
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
